"""
RTBIO 견적서 — Option A / Option B 각각 엑셀 원페이지 생성
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
import os

DIR = os.path.dirname(os.path.abspath(__file__))

# ── Colors ──
PRIMARY = "1B3A5C"
ACCENT = "00A8B5"
WHITE = "FFFFFF"
LIGHT_BG = "F2F6FA"
BORDER_CLR = "D0D5DD"
GRAY = "6B7280"
BLACK = "1A1A2E"
HIGHLIGHT_BG = "E8EEF5"
TOTAL_BG = "1B3A5C"

thin_border = Border(
    left=Side(style="thin", color=BORDER_CLR),
    right=Side(style="thin", color=BORDER_CLR),
    top=Side(style="thin", color=BORDER_CLR),
    bottom=Side(style="thin", color=BORDER_CLR),
)
no_border = Border()
bottom_border = Border(bottom=Side(style="medium", color=PRIMARY))

font_title = Font(name="맑은 고딕", size=18, bold=True, color=PRIMARY)
font_subtitle = Font(name="맑은 고딕", size=10, color=GRAY)
font_section = Font(name="맑은 고딕", size=11, bold=True, color=WHITE)
font_header = Font(name="맑은 고딕", size=9, bold=True, color=PRIMARY)
font_normal = Font(name="맑은 고딕", size=9, color=BLACK)
font_bold = Font(name="맑은 고딕", size=9, bold=True, color=BLACK)
font_gray = Font(name="맑은 고딕", size=9, color=GRAY)
font_total_label = Font(name="맑은 고딕", size=10, bold=True, color=WHITE)
font_total_value = Font(name="맑은 고딕", size=10, bold=True, color=WHITE)
font_note = Font(name="맑은 고딕", size=8, color=GRAY)
font_accent = Font(name="맑은 고딕", size=9, bold=True, color=ACCENT)
font_logo = Font(name="맑은 고딕", size=14, bold=True, color=PRIMARY)

fill_section = PatternFill("solid", fgColor=PRIMARY)
fill_light = PatternFill("solid", fgColor=LIGHT_BG)
fill_highlight = PatternFill("solid", fgColor=HIGHLIGHT_BG)
fill_total = PatternFill("solid", fgColor=TOTAL_BG)
fill_white = PatternFill("solid", fgColor=WHITE)
fill_accent_light = PatternFill("solid", fgColor="E0F7FA")

align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)


def set_cell(ws, row, col, value, font=font_normal, fill=None, alignment=align_left, border=thin_border, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    if fill:
        c.fill = fill
    c.alignment = alignment
    c.border = border
    if number_format:
        c.number_format = number_format
    return c


def section_row(ws, row, text, col_start=1, col_end=6):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill_section
        cell.border = thin_border
    set_cell(ws, row, col_start, text, font=font_section, fill=fill_section, alignment=align_left)
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)


def meta_row(ws, row, label, value, col_label=1, col_value=2, col_end=3):
    set_cell(ws, row, col_label, label, font=font_gray, fill=fill_white, alignment=align_left)
    set_cell(ws, row, col_value, value, font=font_bold, fill=fill_white, alignment=align_left)
    if col_end > col_value:
        ws.merge_cells(start_row=row, start_column=col_value, end_row=row, end_column=col_end)
        for c in range(col_value, col_end + 1):
            ws.cell(row=row, column=c).border = thin_border


def item_row(ws, row, no, item, spec, qty, unit_price, amount, fill=None, is_total=False):
    f = fill or fill_white
    fn = font_bold if is_total else font_normal
    fn_r = font_bold if is_total else font_normal
    set_cell(ws, row, 1, no, font=fn, fill=f, alignment=align_center)
    set_cell(ws, row, 2, item, font=fn, fill=f, alignment=align_left)
    set_cell(ws, row, 3, spec, font=font_gray if not is_total else font_bold, fill=f, alignment=align_left)
    set_cell(ws, row, 4, qty, font=fn, fill=f, alignment=align_center)
    set_cell(ws, row, 5, unit_price, font=fn_r, fill=f, alignment=align_right, number_format='#,##0')
    set_cell(ws, row, 6, amount, font=fn_r, fill=f, alignment=align_right, number_format='#,##0')


def build_option(option):
    wb = Workbook()
    ws = wb.active
    ws.title = f"견적서_Option_{option}"

    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.orientation = "portrait"
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14

    r = 1

    # ── Header ──
    set_cell(ws, r, 1, "HAMADA LABS", font=font_logo, fill=None, alignment=align_left, border=no_border)
    option_label = "Option A — 일시불 모델" if option == "A" else "Option B — 구독형 모델"
    set_cell(ws, r, 5, option_label, font=Font(name="맑은 고딕", size=9, bold=True, color=ACCENT), fill=None, alignment=align_right, border=no_border)
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
    r += 1

    set_cell(ws, r, 1, "견 적 서", font=font_title, fill=None, alignment=align_left, border=no_border)
    r += 1
    set_cell(ws, r, 1, "RTBIO 업무 자동화 시스템", font=font_subtitle, fill=None, alignment=align_left, border=no_border)
    r += 2

    # ── Meta Info ──
    section_row(ws, r, "문서 정보")
    r += 1

    # Left: 공급자, Right: 수신
    meta_pairs = [
        ("공급자", "", "수신", ""),
        ("상호", "하마다랩스", "상호", "알티바이오"),
        ("대표", "방승애", "견적일", "2026년 4월 10일"),
        ("연락처", "010-8888-0180", "유효기간", "발행일로부터 30일"),
        ("이메일", "victoria@hamadalabs.com", "", ""),
    ]
    for lbl_l, val_l, lbl_r, val_r in meta_pairs:
        if lbl_l in ("공급자", "수신"):
            set_cell(ws, r, 1, lbl_l, font=font_accent, fill=fill_light, alignment=align_left)
            set_cell(ws, r, 2, "", font=font_normal, fill=fill_light, alignment=align_left)
            set_cell(ws, r, 3, "", font=font_normal, fill=fill_light, alignment=align_left)
            set_cell(ws, r, 4, lbl_r, font=font_accent, fill=fill_light, alignment=align_left)
            set_cell(ws, r, 5, "", font=font_normal, fill=fill_light, alignment=align_left)
            set_cell(ws, r, 6, "", font=font_normal, fill=fill_light, alignment=align_left)
        else:
            set_cell(ws, r, 1, lbl_l, font=font_gray, fill=fill_white, alignment=align_left)
            set_cell(ws, r, 2, val_l, font=font_bold, fill=fill_white, alignment=align_left)
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
            for c in [2, 3]:
                ws.cell(row=r, column=c).border = thin_border
            set_cell(ws, r, 4, lbl_r, font=font_gray, fill=fill_white, alignment=align_left)
            set_cell(ws, r, 5, val_r, font=font_bold, fill=fill_white, alignment=align_left)
            ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
            for c in [5, 6]:
                ws.cell(row=r, column=c).border = thin_border
        r += 1
    r += 1

    # ── 견적 금액 ──
    if option == "A":
        total_amount = 20000000
        total_label = "견적 총액 (개발비)"
    else:
        total_amount = 10000000
        total_label = "견적 총액 (초기 개발비)"

    section_row(ws, r, "견적 금액")
    r += 1
    # Header
    headers = ["No.", "항목", "상세", "수량", "단가 (원)", "금액 (원)"]
    for i, h in enumerate(headers, 1):
        al = align_center if i in (1, 4) else (align_right if i in (5, 6) else align_left)
        set_cell(ws, r, i, h, font=font_header, fill=fill_highlight, alignment=al)
    r += 1

    if option == "A":
        items = [
            (1, "착수금", "계약 체결 시 (개발비 50%)", 1, 10000000, 10000000),
            (2, "잔금", "최종 납품 완료 시 (개발비 50%)", 1, 10000000, 10000000),
        ]
        for no, item, spec, qty, up, amt in items:
            item_row(ws, r, no, item, spec, qty, up, amt, fill=fill_white)
            r += 1

        # Total
        for c in range(1, 7):
            ws.cell(row=r, column=c).fill = fill_total
            ws.cell(row=r, column=c).border = thin_border
        set_cell(ws, r, 1, "", font=font_total_label, fill=fill_total, alignment=align_center)
        set_cell(ws, r, 2, total_label, font=font_total_label, fill=fill_total, alignment=align_left)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        for c in [2, 3, 4]:
            ws.cell(row=r, column=c).border = thin_border
            ws.cell(row=r, column=c).fill = fill_total
        set_cell(ws, r, 5, "", font=font_total_value, fill=fill_total, alignment=align_right)
        set_cell(ws, r, 6, f'=SUM(F{r-2}:F{r-1})', font=font_total_value, fill=fill_total, alignment=align_right, number_format='#,##0')
        r += 2

        # Monthly
        section_row(ws, r, "월 운영비 (납품 후)")
        r += 1
        for i, h in enumerate(headers, 1):
            al = align_center if i in (1, 4) else (align_right if i in (5, 6) else align_left)
            set_cell(ws, r, i, h, font=font_header, fill=fill_highlight, alignment=al)
        r += 1
        item_row(ws, r, 1, "무상 유지보수", "최종 납품 후 3개월", 3, 0, 0, fill=fill_white)
        r += 1
        item_row(ws, r, 2, "인프라 + 유지보수", "4개월차부터 (인프라 운영, 버그 수정, 보안 패치)", 1, 200000, 200000, fill=fill_white)
        r += 1
        item_row(ws, r, 3, "AI/LLM 사용료", "실비 별도 (1차 납품 후 측정)", 1, None, None, fill=fill_white)
        set_cell(ws, r, 5, "실비", font=font_gray, fill=fill_white, alignment=align_right)
        set_cell(ws, r, 6, "실비", font=font_gray, fill=fill_white, alignment=align_right)
        r += 1

        # Monthly total
        for c in range(1, 7):
            ws.cell(row=r, column=c).fill = fill_total
            ws.cell(row=r, column=c).border = thin_border
        set_cell(ws, r, 1, "", font=font_total_label, fill=fill_total, alignment=align_center)
        set_cell(ws, r, 2, "월 운영비 (무상 기간 이후)", font=font_total_label, fill=fill_total, alignment=align_left)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        for c in [2, 3, 4]:
            ws.cell(row=r, column=c).border = thin_border
            ws.cell(row=r, column=c).fill = fill_total
        set_cell(ws, r, 5, "", font=font_total_value, fill=fill_total)
        set_cell(ws, r, 6, "월 20만원 + AI 실비", font=Font(name="맑은 고딕", size=9, bold=True, color=WHITE), fill=fill_total, alignment=align_right)
        r += 2

    else:  # Option B
        items = [
            (1, "착수금", "계약 체결 시 (초기 개발비 50%)", 1, 5000000, 5000000),
            (2, "잔금", "1차 납품 완료 시 (초기 개발비 50%)", 1, 5000000, 5000000),
        ]
        for no, item, spec, qty, up, amt in items:
            item_row(ws, r, no, item, spec, qty, up, amt, fill=fill_white)
            r += 1

        # Total
        for c in range(1, 7):
            ws.cell(row=r, column=c).fill = fill_total
            ws.cell(row=r, column=c).border = thin_border
        set_cell(ws, r, 1, "", font=font_total_label, fill=fill_total, alignment=align_center)
        set_cell(ws, r, 2, total_label, font=font_total_label, fill=fill_total, alignment=align_left)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        for c in [2, 3, 4]:
            ws.cell(row=r, column=c).border = thin_border
            ws.cell(row=r, column=c).fill = fill_total
        set_cell(ws, r, 5, "", font=font_total_value, fill=fill_total, alignment=align_right)
        set_cell(ws, r, 6, f'=SUM(F{r-2}:F{r-1})', font=font_total_value, fill=fill_total, alignment=align_right, number_format='#,##0')
        r += 2

        # Monthly
        section_row(ws, r, "월 구독료 (36개월 약정)")
        r += 1
        for i, h in enumerate(headers, 1):
            al = align_center if i in (1, 4) else (align_right if i in (5, 6) else align_left)
            set_cell(ws, r, i, h, font=font_header, fill=fill_highlight, alignment=al)
        r += 1
        item_row(ws, r, 1, "월 구독료", "인프라 운영 + 유지보수 + 기능 업데이트 포함", 36, 580000, 580000, fill=fill_white)
        r += 1
        item_row(ws, r, 2, "AI/LLM 사용료", "실비 별도 (1차 납품 후 측정)", 1, None, None, fill=fill_white)
        set_cell(ws, r, 5, "실비", font=font_gray, fill=fill_white, alignment=align_right)
        set_cell(ws, r, 6, "실비", font=font_gray, fill=fill_white, alignment=align_right)
        r += 1

        # Monthly total
        for c in range(1, 7):
            ws.cell(row=r, column=c).fill = fill_total
            ws.cell(row=r, column=c).border = thin_border
        set_cell(ws, r, 1, "", font=font_total_label, fill=fill_total, alignment=align_center)
        set_cell(ws, r, 2, "월 구독료", font=font_total_label, fill=fill_total, alignment=align_left)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        for c in [2, 3, 4]:
            ws.cell(row=r, column=c).border = thin_border
            ws.cell(row=r, column=c).fill = fill_total
        set_cell(ws, r, 5, "", font=font_total_value, fill=fill_total)
        set_cell(ws, r, 6, "월 58만원 + AI 실비", font=Font(name="맑은 고딕", size=9, bold=True, color=WHITE), fill=fill_total, alignment=align_right)
        r += 2

    # ── 제공 범위 ──
    section_row(ws, r, "제공 범위 (1~3차 통합, 2개월 이내 납품)")
    r += 1
    scope_headers = ["No.", "단계", "기능", "", "비고", ""]
    for i, h in enumerate(scope_headers, 1):
        set_cell(ws, r, i, h, font=font_header, fill=fill_highlight, alignment=align_center if i == 1 else align_left)
    r += 1

    scopes = [
        ("1차", "경영지원팀 자동화", "거래처 발주 포털, 거래처/품목 관리, 발주 접수·처리", "메인 시스템 + DB 인프라 구축 포함"),
        ("", "", "거래명세서 자동 발행, 정산 관리, 월 마감", ""),
        ("", "", "UDI 관리, 권한/계정 관리, 어드민 대시보드", ""),
        ("2차", "품질관리팀 재고", "재고 자동 반영, 입고/출고 관리", "1차 인프라 위에 모듈 추가"),
        ("", "", "재고 현황 대시보드, 일별 재고 보고", ""),
        ("", "", "담당자 배정, 작업 상태 관리, 샘플/미팅 출고", ""),
        ("3차", "영업부 대시보드", "거래처별 매출·미수금 현황, 결제 조건 조회", "권한 분리형 대시보드"),
        ("", "", "영업 대시보드, 사용량 입력, 기간별 매출 보고서", ""),
    ]
    no = 0
    for phase, name, funcs, note in scopes:
        f = fill_white if no % 2 == 0 else fill_light
        if phase:
            no += 1
        set_cell(ws, r, 1, no if phase else "", font=font_bold if phase else font_normal, fill=f, alignment=align_center)
        set_cell(ws, r, 2, name, font=font_bold if phase else font_normal, fill=f, alignment=align_left)
        set_cell(ws, r, 3, funcs, font=font_normal, fill=f, alignment=align_left)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        for c in [3, 4]:
            ws.cell(row=r, column=c).border = thin_border
        set_cell(ws, r, 5, note, font=font_gray, fill=f, alignment=align_left)
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
        for c in [5, 6]:
            ws.cell(row=r, column=c).border = thin_border
        r += 1
    r += 1

    # ── 비고 / 특이사항 ──
    section_row(ws, r, "비고")
    r += 1
    notes = [
        "• 세부 기능은 2차 미팅 완료 후 최종 확정합니다.",
        "• 확정 범위 내 자잘한 기능 조정은 무상 반영, 신규 모듈 등 대규모 기능 추가는 별도 견적으로 진행합니다.",
        "• 얼마에요 엑셀 데이터(약 3년치) 이관 비용은 개발비에 포함됩니다.",
        "• AI/LLM 사용료는 1차 납품 후 실제 사용량 측정하여 안내드리며, 실비 기준 별도 청구됩니다.",
        "• 클라우드 기반(AWS) 웹 서비스로 구축하며, 앱 설치 없이 모바일 반응형 웹으로 운영합니다.",
    ]
    if option == "A":
        notes.append("• 개발비 완납 시 소스코드 및 시스템 소유권이 이전됩니다.")
    else:
        notes.append("• 구독 기간(36개월) 중 사용권이 부여되며, 완료 후 소유권이 이전됩니다.")
        notes.append("• 중도 해지 시 잔여 구독료의 30%를 위약금으로 정산합니다.")

    for note in notes:
        set_cell(ws, r, 1, note, font=font_note, fill=fill_white, alignment=align_left, border=thin_border)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = thin_border
        ws.row_dimensions[r].height = 18
        r += 1

    r += 1
    # ── Footer ──
    set_cell(ws, r, 1, "© 2026 HAMADA LABS", font=Font(name="맑은 고딕", size=8, color=GRAY), fill=None, alignment=align_left, border=no_border)

    # Row heights
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 16

    fname = f"RTBIO_견적서_Option_{option}.xlsx"
    out = os.path.join(DIR, fname)
    wb.save(out)
    print(f"✓ {fname} → {out}")


build_option("A")
build_option("B")
print("\nDone!")
